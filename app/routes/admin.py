from flask import Blueprint, render_template, request, flash, redirect, url_for
from flask_login import login_required, current_user
from app import db
from app.models import User, Enrollment, StudentSchedule, Subject, TimeSlot, TeacherAvailability, Program, Batch, ProgramSubject, TeacherSkill, Booking, Attendance, Tool, ProgramTool
from datetime import date

bp = Blueprint('admin', __name__, url_prefix='/admin')

# Decorator
def admin_required(func):
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Akses ditolak!')
            return redirect(url_for('main.dashboard'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

# --- SISWA (STUDENT) ---
@bp.route('/students')
@login_required
@admin_required
def student_list():
    students = db.session.query(User).filter(User.role == 'student').all()
    return render_template('admin/student_list.html', students=students)

@bp.route('/student/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def student_detail(user_id):
    student = User.query.get_or_404(user_id)
    enrollment = Enrollment.query.filter_by(student_id=student.id).first()
    
    if request.method == 'POST':
        if 'update_info' in request.form:
            enrollment.sessions_remaining = int(request.form['sessions_remaining'])
            enrollment.status = request.form['status']
            db.session.commit()
            flash('Info siswa diperbarui.')
        
        elif 'add_schedule' in request.form:
            new_sched = StudentSchedule(
                enrollment_id=enrollment.id,
                day_of_week=int(request.form['day']),
                timeslot_id=int(request.form['timeslot_id']),
                subject_id=int(request.form['subject_id']),
                teacher_id=int(request.form['teacher_id'])
            )
            db.session.add(new_sched)
            db.session.commit()
            flash('Jadwal manual ditambahkan.')
        
        elif 'add_manual_booking' in request.form:
            booking_date_str = request.form['date']
            booking_date = date.fromisoformat(booking_date_str)
            timeslot_id = int(request.form['timeslot_id'])
            
            # Additional fields for teacher/subject
            teacher_id = request.form.get('teacher_id')
            subject_id = request.form.get('subject_id')
            
            if teacher_id: teacher_id = int(teacher_id)
            if subject_id: subject_id = int(subject_id)
            
            # Check duplicate
            existing = Booking.query.filter_by(
                enrollment_id=enrollment.id, 
                date=booking_date, 
                timeslot_id=timeslot_id
            ).first()
            
            if existing:
                flash('Booking untuk tanggal dan jam tersebut sudah ada!', 'error')
            else:
                new_booking = Booking(
                    enrollment_id=enrollment.id,
                    date=booking_date,
                    timeslot_id=timeslot_id,
                    teacher_id=teacher_id, # Added
                    subject_id=subject_id, # Added
                    status='booked'
                )
                db.session.add(new_booking)
                db.session.commit()
                flash('Override jadwal (Sesi Tambahan) berhasil dibuat.')

        return redirect(url_for('admin.student_detail', user_id=user_id))

    booking_date_cutoff = date.today()
    manual_bookings = Booking.query.filter(
        Booking.enrollment_id == enrollment.id,
        Booking.status != 'completed',
        Booking.date >= booking_date_cutoff
    ).order_by(Booking.date).all()

    # --- STUDENT PROGRESS DATA ---
    student_progress = None
    if enrollment:
        # Get completed bookings
        completed_bookings = Booking.query.filter(
            Booking.enrollment_id == enrollment.id,
            Booking.status == 'completed'
        ).all()
        booking_ids = [b.id for b in completed_bookings]
        
        # Attendance Stats
        hadir_count = Attendance.query.filter(Attendance.booking_id.in_(booking_ids), Attendance.status == 'Hadir').count() if booking_ids else 0
        izin_count = Attendance.query.filter(Attendance.booking_id.in_(booking_ids), Attendance.status == 'Izin').count() if booking_ids else 0
        alpha_count = Attendance.query.filter(Attendance.booking_id.in_(booking_ids), Attendance.status == 'Alpha').count() if booking_ids else 0
        
        # Progress Pct
        total_sessions = enrollment.program.total_sessions if enrollment.program else 0
        completed_sessions_count = total_sessions - enrollment.sessions_remaining
        progress_pct = int((completed_sessions_count / total_sessions) * 100) if total_sessions > 0 else 0
        
        # Session History
        attendance_records = Attendance.query.filter(
            Attendance.booking_id.in_(booking_ids)
        ).order_by(Attendance.date.desc()).limit(10).all() if booking_ids else []
        
        session_history = []
        for att in attendance_records:
            booking = Booking.query.get(att.booking_id)
            if booking:
                session_history.append({
                    'date': att.date,
                    'subject': booking.subject.name if booking.subject else '-',
                    'timeslot': booking.timeslot.name if booking.timeslot else '-',
                    'teacher': att.teacher.name if att.teacher else '-',
                    'status': att.status,
                    'notes': att.notes
                })

        student_progress = {
            'completed_sessions': completed_sessions_count,
            'total_sessions': total_sessions,
            'remaining_sessions': enrollment.sessions_remaining,
            'progress_pct': progress_pct,
            'hadir': hadir_count,
            'izin': izin_count,
            'alpha': alpha_count,
            'session_history': session_history
        }

    subjects = Subject.query.all()
    timeslots = TimeSlot.query.all()
    teachers = User.query.filter_by(role='teacher').all()
    days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']

    return render_template('admin/student_detail.html', 
                           student=student, 
                           enrollment=enrollment, 
                           subjects=subjects, 
                           timeslots=timeslots, 
                           teachers=teachers, 
                           days=days, 
                           manual_bookings=manual_bookings,
                           student_progress=student_progress)

@bp.route('/booking/delete/<int:booking_id>', methods=['POST'])
@login_required
@admin_required
def delete_booking(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    uid = booking.enrollment.student_id
    db.session.delete(booking)
    db.session.commit()
    flash('Booking manual dihapus.')
    return redirect(url_for('admin.student_detail', user_id=uid))

@bp.route('/schedule/delete/<int:sched_id>')
@login_required
@admin_required
def delete_schedule(sched_id):
    sched = StudentSchedule.query.get_or_404(sched_id)
    uid = sched.enrollment.student_id
    db.session.delete(sched)
    db.session.commit()
    flash('Jadwal dihapus.')
    return redirect(url_for('admin.student_detail', user_id=uid))

@bp.route('/master-schedule')
@login_required
@admin_required
def master_schedule():
    today_date = date.today()
    days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    timeslots = TimeSlot.query.all()
    
    # Initialize aggregated stats
    total_students = db.session.query(User).filter(User.role == 'student').count()
    total_schedules = 0
    day_totals = {i: 0 for i in range(7)}
    
    schedule_map = {i: {slot.id: [] for slot in timeslots} for i in range(7)}
    
    # Query all schedules is heavier, better filter per slot via loop or join
    # Simple loop approach:
    for i in range(7):
        for slot in timeslots:
            schedules = StudentSchedule.query.filter_by(day_of_week=i, timeslot_id=slot.id).all()
            schedule_map[i][slot.id] = schedules
            count = len(schedules)
            total_schedules += count
            day_totals[i] += count
            
    return render_template('admin/master_schedule.html', 
                           days=days, 
                           timeslots=timeslots, 
                           schedule_map=schedule_map,
                           total_students=total_students,
                           total_schedules=total_schedules,
                           day_totals=day_totals)

# --- PROGRAM ---
@bp.route('/programs', methods=['GET', 'POST'])
@login_required
@admin_required
def program_manage():
    if request.method == 'POST':
        name = request.form['name']
        sessions = int(request.form['total_sessions'])
        is_batch = True if request.form.get('is_batch_based') else False
        
        new_prog = Program(name=name, total_sessions=sessions, is_batch_based=is_batch)
        db.session.add(new_prog)
        db.session.commit()
        
        # Add Subjects
        sub_ids = request.form.getlist('subjects')
        for sid in sub_ids:
            db.session.add(ProgramSubject(program_id=new_prog.id, subject_id=int(sid)))
        db.session.commit()
        flash('Program ditambahkan.')
        return redirect(url_for('admin.program_manage'))

    programs = Program.query.all()
    subjects = Subject.query.all()
    return render_template('admin/program_list.html', programs=programs, subjects=subjects)

@bp.route('/program/edit/<int:prog_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def program_edit(prog_id):
    prog = Program.query.get_or_404(prog_id)
    if request.method == 'POST':
        prog.name = request.form['name']
        prog.total_sessions = int(request.form['total_sessions'])
        prog.is_batch_based = True if request.form.get('is_batch_based') else False
        
        ProgramSubject.query.filter_by(program_id=prog.id).delete()
        for sid in request.form.getlist('subjects'):
            db.session.add(ProgramSubject(program_id=prog.id, subject_id=int(sid)))
        db.session.commit()
        flash('Program diupdate.')
        return redirect(url_for('admin.program_manage'))

    current_sub_ids = [ps.subject_id for ps in prog.subjects]
    return render_template('admin/program_edit.html', program=prog, subjects=Subject.query.all(), current_sub_ids=current_sub_ids)

# --- BATCH MANAGEMENT ---
@bp.route('/batch/add', methods=['POST'])
@login_required
@admin_required
def add_batch():
    program_id = int(request.form['program_id'])
    name = request.form['name']
    max_students = int(request.form['max_students'])
    
    new_batch = Batch(program_id=program_id, name=name, max_students=max_students, is_active=True)
    db.session.add(new_batch)
    db.session.commit()
    flash('Batch berhasil ditambahkan.')
    return redirect(url_for('admin.program_edit', prog_id=program_id))

@bp.route('/batch/edit/<int:batch_id>', methods=['POST'])
@login_required
@admin_required
def edit_batch(batch_id):
    batch = Batch.query.get_or_404(batch_id)
    batch.name = request.form['name']
    batch.max_students = int(request.form['max_students'])
    batch.is_active = True if request.form.get('is_active') else False
    
    db.session.commit()
    flash('Batch diperbarui.')
    return redirect(url_for('admin.program_edit', prog_id=batch.program_id))

@bp.route('/batch/delete/<int:batch_id>', methods=['POST'])
@login_required
@admin_required
def delete_batch(batch_id):
    batch = Batch.query.get_or_404(batch_id)
    program_id = batch.program_id
    
    # Check for enrolled students
    if Enrollment.query.filter_by(batch_id=batch.id).first():
        flash('Gagal menghapus: Masih ada siswa yang terdaftar di batch ini.', 'error')
        return redirect(url_for('admin.program_edit', prog_id=program_id))
        
    try:
        db.session.delete(batch)
        db.session.commit()
        flash('Batch dihapus.')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus batch: {str(e)}', 'error')
        
    return redirect(url_for('admin.program_edit', prog_id=program_id))

# --- TEACHER ---
@bp.route('/teachers', methods=['GET', 'POST'])
@login_required
@admin_required
def teacher_list():
    if request.method == 'POST':
        email = request.form['email']
        # Cek duplicate
        if User.query.filter_by(email=email).first():
            flash('Email sudah ada.')
        else:
            new_t = User(
                email=email, 
                name=request.form['name'], 
                phone_number=request.form['phone'], 
                role='teacher'
            )
            # PASSWORD DEFAULT
            new_t.set_password('guru123')
            db.session.add(new_t)
            db.session.commit()
            flash('Guru ditambahkan. Password default: guru123')
            return redirect(url_for('admin.teacher_list'))

    teachers = User.query.filter_by(role='teacher').all()
    return render_template('admin/teacher_list.html', teachers=teachers)

@bp.route('/teacher/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def teacher_detail(user_id):
    teacher = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        if 'update_profile' in request.form:
            teacher.name = request.form['name']
            teacher.email = request.form['email']
            teacher.phone_number = request.form['phone']
            db.session.commit()
            flash('Profil diupdate.')
            
        elif 'update_skills' in request.form:
            TeacherSkill.query.filter_by(teacher_id=teacher.id).delete()
            for sid in request.form.getlist('subject_ids'):
                db.session.add(TeacherSkill(teacher_id=teacher.id, subject_id=int(sid)))
            db.session.commit()
            flash('Skill diupdate.')
            
        elif 'update_avail' in request.form:
            TeacherAvailability.query.filter_by(teacher_id=teacher.id).delete()
            for item in request.form.getlist('slots'):
                day, slot = item.split('_')
                db.session.add(TeacherAvailability(teacher_id=teacher.id, day_of_week=int(day), timeslot_id=int(slot)))
            db.session.commit()
            flash('Jadwal ketersediaan diupdate.')
            
        return redirect(url_for('admin.teacher_detail', user_id=user_id))

    subjects = Subject.query.all()
    timeslots = TimeSlot.query.all()
    days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    
    my_skill_ids = [s.subject_id for s in teacher.skills]
    avail_map = {i: {t.id: False for t in timeslots} for i in range(7)}
    for av in teacher.availabilities:
        avail_map[av.day_of_week][av.timeslot_id] = True
        
    return render_template('admin/teacher_detail.html', teacher=teacher, subjects=subjects, timeslots=timeslots, days=days, my_skill_ids=my_skill_ids, avail_map=avail_map)

# --- DELETE ROUTES ---
@bp.route('/student/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_student(user_id):
    student = User.query.get_or_404(user_id)
    if student.role != 'student':
        flash('User bukan siswa.')
        return redirect(url_for('admin.student_list'))
        
    try:
        db.session.delete(student)
        db.session.commit()
        flash('Siswa dihapus.')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus: {str(e)}')
        
    return redirect(url_for('admin.student_list'))

@bp.route('/program/delete/<int:prog_id>', methods=['POST'])
@login_required
@admin_required
def delete_program(prog_id):
    prog = Program.query.get_or_404(prog_id)
    try:
        db.session.delete(prog)
        db.session.commit()
        flash('Program dihapus.')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus program: {str(e)}')
        
    return redirect(url_for('admin.program_manage'))

@bp.route('/teacher/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_teacher(user_id):
    teacher = User.query.get_or_404(user_id)
    if teacher.role != 'teacher':
        flash('User bukan guru.')
        return redirect(url_for('admin.teacher_list'))
        
    try:
        db.session.delete(teacher)
        db.session.commit()
        flash('Guru dihapus.')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus guru: {str(e)}')
        
    return redirect(url_for('admin.teacher_list'))

# --- TEACHER RECAP ---
@bp.route('/teacher-recap')
@login_required
@admin_required
def teacher_recap():
    """Display teacher session recap page"""
    teachers = User.query.filter_by(role='teacher').all()
    selected_teacher_id = request.args.get('teacher_id', type=int)
    selected_year = request.args.get('year', type=int, default=date.today().year)
    
    recap_data = None
    selected_teacher = None
    
    if selected_teacher_id:
        selected_teacher = User.query.get(selected_teacher_id)
        if selected_teacher:
            # Get all attendance records where this teacher conducted the session
            recap_data = []
            for month in range(1, 13):
                # Count unique bookings (sessions) for this teacher in this month
                # A session is counted once regardless of how many students attended
                from sqlalchemy import extract, func
                
                session_count = db.session.query(func.count(func.distinct(Attendance.booking_id))).filter(
                    Attendance.teacher_id == selected_teacher_id,
                    extract('year', Attendance.date) == selected_year,
                    extract('month', Attendance.date) == month
                ).scalar() or 0
                
                recap_data.append({
                    'month': month,
                    'month_name': ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'][month-1],
                    'sessions': session_count
                })
    
    # Available years for filter
    years = list(range(2024, date.today().year + 2))
    
    return render_template('admin/teacher_recap.html',
                          teachers=teachers,
                          selected_teacher=selected_teacher,
                          selected_year=selected_year,
                          recap_data=recap_data,
                          years=years)

@bp.route('/teacher-recap/export-excel')
@login_required
@admin_required
def export_recap_excel():
    """Export teacher recap to Excel"""
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    teacher_id = request.args.get('teacher_id', type=int)
    year = request.args.get('year', type=int, default=date.today().year)
    
    if not teacher_id:
        flash('Pilih pengajar terlebih dahulu.')
        return redirect(url_for('admin.teacher_recap'))
    
    teacher = User.query.get(teacher_id)
    if not teacher:
        flash('Pengajar tidak ditemukan.')
        return redirect(url_for('admin.teacher_recap'))
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rekap {year}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Rekap Sesi Pengajar: {teacher.name}"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws.merge_cells('A2:C2')
    ws['A2'] = f"Tahun: {year}"
    ws['A2'].font = Font(size=12)
    
    # Headers
    headers = ['Bulan', 'Jumlah Sesi', 'Keterangan']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
              'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    
    total_sessions = 0
    from sqlalchemy import extract, func
    
    for row, (month_num, month_name) in enumerate(enumerate(months, 1), 5):
        session_count = db.session.query(func.count(func.distinct(Attendance.booking_id))).filter(
            Attendance.teacher_id == teacher_id,
            extract('year', Attendance.date) == year,
            extract('month', Attendance.date) == month_num
        ).scalar() or 0
        
        total_sessions += session_count
        
        ws.cell(row=row, column=1, value=month_name).border = thin_border
        ws.cell(row=row, column=2, value=session_count).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value='-').border = thin_border
    
    # Total row
    total_row = 17
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border
    ws.cell(row=total_row, column=2, value=total_sessions).font = Font(bold=True)
    ws.cell(row=total_row, column=2).border = thin_border
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=3, value='').border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rekap_{teacher.name.replace(' ', '_')}_{year}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@bp.route('/teacher-recap/export-pdf')
@login_required
@admin_required
def export_recap_pdf():
    """Export teacher recap to PDF"""
    from io import BytesIO
    from flask import send_file
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    
    teacher_id = request.args.get('teacher_id', type=int)
    year = request.args.get('year', type=int, default=date.today().year)
    
    if not teacher_id:
        flash('Pilih pengajar terlebih dahulu.')
        return redirect(url_for('admin.teacher_recap'))
    
    teacher = User.query.get(teacher_id)
    if not teacher:
        flash('Pengajar tidak ditemukan.')
        return redirect(url_for('admin.teacher_recap'))
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, spaceAfter=20)
    elements.append(Paragraph(f"Rekap Sesi Pengajar", title_style))
    elements.append(Paragraph(f"<b>Nama:</b> {teacher.name}", styles['Normal']))
    elements.append(Paragraph(f"<b>Tahun:</b> {year}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Bulan', 'Jumlah Sesi']]
    months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
              'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    
    total_sessions = 0
    from sqlalchemy import extract, func
    
    for month_num, month_name in enumerate(months, 1):
        session_count = db.session.query(func.count(func.distinct(Attendance.booking_id))).filter(
            Attendance.teacher_id == teacher_id,
            extract('year', Attendance.date) == year,
            extract('month', Attendance.date) == month_num
        ).scalar() or 0
        
        total_sessions += session_count
        data.append([month_name, str(session_count)])
    
    data.append(['TOTAL', str(total_sessions)])
    
    # Create table
    table = Table(data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D32F2F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFEBEE')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rekap_{teacher.name.replace(' ', '_')}_{year}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

# --- EXPORT ALL TEACHERS ---
@bp.route('/teacher-recap/export-all-monthly-excel')
@login_required
@admin_required
def export_all_monthly_excel():
    """Export all teachers monthly recap to Excel"""
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from sqlalchemy import extract, func
    
    year = request.args.get('year', type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)
    
    teachers = User.query.filter_by(role='teacher').all()
    months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
              'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rekap {months[month-1]} {year}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Rekap Sesi Semua Pengajar - {months[month-1]} {year}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ['No', 'Nama Pengajar', 'Jumlah Sesi']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    total_all = 0
    for idx, teacher in enumerate(teachers, 1):
        session_count = db.session.query(func.count(func.distinct(Attendance.booking_id))).filter(
            Attendance.teacher_id == teacher.id,
            extract('year', Attendance.date) == year,
            extract('month', Attendance.date) == month
        ).scalar() or 0
        
        total_all += session_count
        row = idx + 3
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=teacher.name).border = thin_border
        ws.cell(row=row, column=3, value=session_count).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    
    # Total row
    total_row = len(teachers) + 4
    ws.cell(row=total_row, column=1, value='').border = thin_border
    ws.cell(row=total_row, column=2, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=2).border = thin_border
    ws.cell(row=total_row, column=3, value=total_all).font = Font(bold=True)
    ws.cell(row=total_row, column=3).border = thin_border
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rekap_semua_pengajar_{months[month-1]}_{year}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@bp.route('/teacher-recap/export-all-yearly-excel')
@login_required
@admin_required
def export_all_yearly_excel():
    """Export all teachers yearly recap to Excel"""
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from sqlalchemy import extract, func
    
    year = request.args.get('year', type=int, default=date.today().year)
    teachers = User.query.filter_by(role='teacher').all()
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Ags', 'Sep', 'Okt', 'Nov', 'Des']
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rekap Tahunan {year}"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:O1')
    ws['A1'] = f"Rekap Sesi Semua Pengajar - Tahun {year}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Headers: No, Nama, Jan, Feb, ..., Des, Total
    headers = ['No', 'Nama Pengajar'] + months + ['TOTAL']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    grand_total = 0
    for idx, teacher in enumerate(teachers, 1):
        row = idx + 3
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=teacher.name).border = thin_border
        
        teacher_total = 0
        for m in range(1, 13):
            session_count = db.session.query(func.count(func.distinct(Attendance.booking_id))).filter(
                Attendance.teacher_id == teacher.id,
                extract('year', Attendance.date) == year,
                extract('month', Attendance.date) == m
            ).scalar() or 0
            
            teacher_total += session_count
            cell = ws.cell(row=row, column=m+2, value=session_count)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        grand_total += teacher_total
        ws.cell(row=row, column=15, value=teacher_total).border = thin_border
        ws.cell(row=row, column=15).font = Font(bold=True)
        ws.cell(row=row, column=15).alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    for col in range(3, 16):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 7
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rekap_semua_pengajar_tahunan_{year}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@bp.route('/teacher-recap/export-all-monthly-pdf')
@login_required
@admin_required
def export_all_monthly_pdf():
    """Export all teachers monthly recap to PDF"""
    from io import BytesIO
    from flask import send_file
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from sqlalchemy import extract, func
    
    year = request.args.get('year', type=int, default=date.today().year)
    month = request.args.get('month', type=int, default=date.today().month)
    
    teachers = User.query.filter_by(role='teacher').all()
    months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
              'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, spaceAfter=20)
    elements.append(Paragraph(f"Rekap Sesi Semua Pengajar", title_style))
    elements.append(Paragraph(f"<b>Periode:</b> {months[month-1]} {year}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['No', 'Nama Pengajar', 'Jumlah Sesi']]
    total_all = 0
    
    for idx, teacher in enumerate(teachers, 1):
        session_count = db.session.query(func.count(func.distinct(Attendance.booking_id))).filter(
            Attendance.teacher_id == teacher.id,
            extract('year', Attendance.date) == year,
            extract('month', Attendance.date) == month
        ).scalar() or 0
        
        total_all += session_count
        data.append([str(idx), teacher.name, str(session_count)])
    
    data.append(['', 'TOTAL', str(total_all)])
    
    table = Table(data, colWidths=[0.5*inch, 3*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D32F2F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFEBEE')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rekap_semua_pengajar_{months[month-1]}_{year}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

@bp.route('/teacher-recap/export-all-yearly-pdf')
@login_required
@admin_required
def export_all_yearly_pdf():
    """Export all teachers yearly recap to PDF"""
    from io import BytesIO
    from flask import send_file
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from sqlalchemy import extract, func
    
    year = request.args.get('year', type=int, default=date.today().year)
    teachers = User.query.filter_by(role='teacher').all()
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Ags', 'Sep', 'Okt', 'Nov', 'Des']
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, spaceAfter=15)
    elements.append(Paragraph(f"Rekap Sesi Semua Pengajar - Tahun {year}", title_style))
    elements.append(Spacer(1, 10))
    
    # Table data
    headers = ['No', 'Nama'] + months + ['Total']
    data = [headers]
    
    for idx, teacher in enumerate(teachers, 1):
        row = [str(idx), teacher.name]
        teacher_total = 0
        for m in range(1, 13):
            session_count = db.session.query(func.count(func.distinct(Attendance.booking_id))).filter(
                Attendance.teacher_id == teacher.id,
                extract('year', Attendance.date) == year,
                extract('month', Attendance.date) == m
            ).scalar() or 0
            teacher_total += session_count
            row.append(str(session_count))
        row.append(str(teacher_total))
        data.append(row)
    
    # Column widths for landscape
    col_widths = [0.4*inch, 1.8*inch] + [0.5*inch]*12 + [0.6*inch]
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D32F2F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rekap_semua_pengajar_tahunan_{year}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

# --- DATA TOOLS ---
@bp.route('/tools')
@login_required
@admin_required
def tools_list():
    """Display list of tools and programs"""
    tools = Tool.query.all()
    programs = Program.query.all()
    return render_template('admin/tools.html', tools=tools, programs=programs)

@bp.route('/tools/add', methods=['POST'])
@login_required
@admin_required
def add_tool():
    """Add a new tool"""
    name = request.form.get('name')
    description = request.form.get('description', '')
    category = request.form.get('category', '')
    
    if name:
        tool = Tool(name=name, description=description, category=category)
        db.session.add(tool)
        db.session.commit()
        flash(f'Tool "{name}" berhasil ditambahkan.')
    else:
        flash('Nama tool wajib diisi.', 'error')
    
    return redirect(url_for('admin.tools_list'))

@bp.route('/tools/delete/<int:tool_id>', methods=['POST'])
@login_required
@admin_required
def delete_tool(tool_id):
    """Delete a tool"""
    tool = Tool.query.get_or_404(tool_id)
    
    # Delete associated program_tools first
    ProgramTool.query.filter_by(tool_id=tool_id).delete()
    
    db.session.delete(tool)
    db.session.commit()
    flash(f'Tool "{tool.name}" berhasil dihapus.')
    return redirect(url_for('admin.tools_list'))

@bp.route('/tools/assign', methods=['POST'])
@login_required
@admin_required
def assign_tool_to_program():
    """Assign a tool to a program"""
    tool_id = request.form.get('tool_id', type=int)
    program_id = request.form.get('program_id', type=int)
    quantity = request.form.get('quantity', type=int, default=1)
    notes = request.form.get('notes', '')
    
    if tool_id and program_id:
        # Check if already assigned
        existing = ProgramTool.query.filter_by(tool_id=tool_id, program_id=program_id).first()
        if existing:
            existing.quantity = quantity
            existing.notes = notes
            flash('Assignment tool diperbarui.')
        else:
            pt = ProgramTool(tool_id=tool_id, program_id=program_id, quantity=quantity, notes=notes)
            db.session.add(pt)
            flash('Tool berhasil ditambahkan ke program.')
        db.session.commit()
    else:
        flash('Pilih tool dan program.', 'error')
    
    return redirect(url_for('admin.tools_list'))

@bp.route('/tools/unassign/<int:pt_id>', methods=['POST'])
@login_required
@admin_required
def unassign_tool(pt_id):
    """Remove tool from program"""
    pt = ProgramTool.query.get_or_404(pt_id)
    db.session.delete(pt)
    db.session.commit()
    flash('Tool dihapus dari program.')
    return redirect(url_for('admin.tools_list'))